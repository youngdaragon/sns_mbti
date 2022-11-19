# %%
import math
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import time
import os
import shutil
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
import persons as pe
import persondetect as ps

# %%
keywords = {
  'E': ['그룹', '모임', '파티', '팸', '축제'],
  'I': ['집', '휴식'],
  'S': ['건설', '솔직', '현실'],
  'N': ['발명', '이념'],
  'T': ['생각', '합리'],
  'F': ['느낌'],
  'J': ['판단'],
  'P': ['유유자적'],
}

# %%
wb = load_workbook('result.xlsx') if os.path.exists('result.xlsx') else Workbook()

# %%
path='/home/kimyongtae/yolov5/chromedriver'
driver = webdriver.Chrome(path)

# %%
driver.get('https://instagram.com')
WebDriverWait(driver, 10).until(
  EC.presence_of_element_located((By.CSS_SELECTOR, "input[name=username]"))
)
driver.find_element(By.CSS_SELECTOR, 'input[name=username]').send_keys(input('username: '))
driver.find_element(By.CSS_SELECTOR, 'input[name=password]').send_keys(input('password: '))
driver.find_element(By.CSS_SELECTOR, 'button[type=submit]').click()

WebDriverWait(driver, 10).until(EC.url_changes(driver.current_url))
if 'two_factor' in driver.current_url:
  otp_code = input('otp code: ')
  driver.find_element(By.CSS_SELECTOR, 'input[type=tel]').send_keys(otp_code)
driver.find_element(By.CSS_SELECTOR, 'button[type=button]').click()

# %%
try:
  WebDriverWait(driver, 10).until(EC.url_to_be('https://www.instagram.com/'))
except:
  WebDriverWait(driver, 10).until(EC.url_contains('onetap'))
  driver.find_element(By.CSS_SELECTOR, 'button').click()
  WebDriverWait(driver, 10).until(EC.url_to_be('https://www.instagram.com/'))

# %%
user_name = input('user name: ')
driver.get(f'https://instagram.com/{user_name}')
WebDriverWait(driver, 10).until(
  EC.presence_of_element_located((By.CSS_SELECTOR, 'article img'))
)

# %%
image_class = driver.find_element(By.CSS_SELECTOR, 'article img').get_attribute('class').strip()
driver.find_element(By.XPATH, '//article//img/../..').click()
WebDriverWait(driver, 10).until(
  EC.presence_of_element_located((By.XPATH, '(//div[@role="dialog"]//button)[5]'))
)

# %%
articles = []

# %%
i = 0
while i >= 0:
  article_image_selector = f'//div[@role="dialog"]//article//*[self::img[@class="{image_class}"] or self::video]'
  WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, article_image_selector))
  )
  pictures = [driver.find_element(By.XPATH, article_image_selector).get_attribute('src')]
  while True:
    if len(driver.find_elements(By.CLASS_NAME, 'coreSpriteRightChevron')) == 0: break
    driver.find_element(By.CLASS_NAME, 'coreSpriteRightChevron').click()
    time.sleep(0.3)
    pictures.append(
      driver.find_elements(By.XPATH, article_image_selector)[1].get_attribute('src')
    )
  articles.append({
    'id': driver.current_url.split('/')[-2],
    'hashtags': [*map(lambda x: x.get_attribute('innerHTML').strip(), driver.find_elements(By.CSS_SELECTOR, 'a[href*="/tags/"]'))],
    'pictures': pictures,
    'username': user_name,
  })
  elems = driver.find_elements(By.XPATH, '//button[@type="button"]//*[local-name()="svg" and @role="img" and @width="16"]/../../..')
  if len(elems) < (1 if i == 0 else 2): break
  elems[(0 if i == 0 else 1)].click()
  i += 1

# %%
count=0
count1=0
for article in articles:
  count+=1
  if not os.path.isdir(article['username']): os.mkdir(article['username'])
  for index in range(len(article['pictures'])):
    count1+=1
    picture = article['pictures'][index]
    if picture.startswith('blob'):
      continue
    r = requests.get(picture, stream=True)
    # shutil.copyfileobj(r.raw, open(f"{article['username']}/{index}.jpg", 'wb'))
    shutil.copyfileobj(r.raw, open("{}/{}+{}+{}.jpg".format(article['username'],article['username'],count,index), 'wb'))
  if count1==50:
    break

# %%
mbti_count = {
  'E': 1,
  'I': 1,
  'S': 1,
  'N': 1,
  'T': 1,
  'F': 1,
  'J': 1,
  'P': 1,
}

hashtags = pd.Series([hashtag for article in articles for hashtag in article['hashtags']])

for mbti, hash_keywords in keywords.items():
  for keyword in hash_keywords:
    mbti_count[mbti] += len(hashtags[hashtags.str.contains(keyword)])

# %%
get_many = lambda a, b: a if mbti_count[a] > mbti_count[b] else b if mbti_count[a] < mbti_count[b] else '?'
result = f"{get_many('E', 'I')}{get_many('S', 'N')}{get_many('T', 'F')}{get_many('J', 'P')}"

# %%
# draw radar chart
categories = [*[*keywords.keys()][::2], *[*keywords.keys()][1::2]]
N = len(categories)

sum_of_all_keywords = sum(mbti_count.values())
mbti_ratio = {k: v / sum_of_all_keywords for k, v in mbti_count.items()}
mbti_ratio = [mbti_ratio[k] for k in categories]
mbti_ratio += mbti_ratio[:1]

angles = [n / float(N) * 2 * math.pi for n in range(N)]
angles += angles[:1]

ax = plt.subplot(111, polar=True)

plt.xticks(angles[:-1], categories, c='grey', size=8)

ax.set_rlabel_position(0)
plt.yticks([0.2, 0.4, 0.6, 0.8], ['20%', '40%', '60%', '80%'])
plt.ylim(0, 1)

ax.plot(angles, mbti_ratio, lw=1, ls='solid')

ax.fill(angles, mbti_ratio, 'b', alpha=0.1)

plt.savefig('chart.png')

# %%
ws = wb.create_sheet(user_name + ' ' + datetime.now().strftime('%Y-%m-%d %H_%M'))

# %%
hashtags_count = hashtags.value_counts()

# %%
hashtags_count.index = hashtags_count.index.map(lambda x: x[1:])

# %%
for tag, count in hashtags_count.iteritems():
  ws.append([tag, count])

# %%
wb.save('result.xlsx')
print(result)

ps.run(weights='yolov5s.pt',source=user_name, save_txt=True, project='result', name='mbti')



# 01091653389
# onit1boss$$!